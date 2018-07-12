import csv
import os.path
import openpyxl as xl
import numpy as np
"""
Consolidata.py

Author: Jordan Kemp
Date: July 12, 2018
For VRDI

Combines the prorated LTSB ward data and Blue Book ward cluster data. Exports to XL
"""

bb = " "
r_file = "ForR.xlsx"
bb_file = "SamCopyAndPastesValues.xlsx"
output_file = "demographics.xlsx"

first_sheet = None

#READING: array constants
REP_PRES = 3
DEM_PRES = 4
COUNTY = 5
TYPE = 6
NAME=7
WARD = 8
ROW = 0

#READING: Ranges of valid LTSB data columns
WSAREP = 61
PREDEM12 = 57
ECO1 = 29
REP_INCUM = 23
DEM_INCUM = 24
CONTESTED = 28
HISP_VEP = 18
TOT_VEP  = 15
HISP_18 = 8
TOT_18 = 6

#WRITING: Columns of output lines
KEY = 1
PREP = 2
PDEM = 3
AREP = 4
ADEM = 5
CTY = 6
ASMB = 7
FIP = 8
DEMOGRAPHICS = 9
ROW1 = ["WardGroup","PREREP12","PREDEM12","WSAREP12","WSADEM12","county","assembly","county_fip"]
ROW_E= None

ARR_INCUM = 9

class District(object):

    @staticmethod
    def append_data(upper,lower,master,NUMS):

        list = []

        if NUMS:
            for i in range(lower,upper):
                try:
                    list.append(np.sum([cell[i].value for cell in master]))
                except:
                    list.append(0)
                    print("excel error")
        else:
            for i in range(lower,upper):
                try:
                    list.append(master[i].value)
                except:
                    print("excel error")
        return list

    def ltsb_append(self,ltsb_list):

        try:
            vote18 = self.append_data(HISP_18+1,TOT_18,ltsb_list,True)
            vep = self.append_data(HISP_VEP+1,TOT_VEP,ltsb_list,True)
            incum = self.append_data(DEM_INCUM+1,REP_INCUM,ltsb_list,True)
            misc = self.append_data(WSAREP+2,ECO1,ltsb_list,True)

            self.ltsb = np.concatenate([vote18, vep, incum, misc],0)
            self.asmbly = ltsb_list[0][2].value

        except:
            self.ltsb = [0 for i in range(30)]


    #District class contains R/D Pres/Asmbly voting data. County, type, Name, and Ward Data
    def __init__(self,row,keys):
        self.rprs = row[REP_PRES].value
        self.dprs = row[DEM_PRES].value
        self.rass = 0
        self.dass = 0
        self.cty = row[COUNTY].value
        self.tp = row[TYPE].value
        self.nm = row[NAME].value.replace(" ","")


        #The wards are input as a list of wards and the key used for the dict
        self.wrd = convert_to_key(row,keys,True)[0]
        self.str = convert_to_key(row,keys,False)[1]

        self.ltsb = None
        self.asmbly = None


#Converts row information into keys for the dictionary of districts
def convert_to_key(row,keys,GENERATING_KEYS):

    #Creates a list of wards by numbers
    w_list = [x.strip() for x in str(row[WARD].value).split(',')]

    #Begins the key with the COUNTY TOWN - TYPE
    key = str(row[COUNTY].value).replace(" ","")+" "+str(row[NAME].value)+" - "+str(row[TYPE].value)

    #Uses this as key for key dictionary -- Maps town to ward combinations
    key_key = key.lower()

    #Appends all ward numbers to individual towns
    key += " "+w_list[0]
    for val in w_list:

        if not val == w_list[0]:
            key += ","+str(val)

    #If this is the first instance of this town, adds to key dictionary
    if GENERATING_KEYS:
        if not key_key in keys:
            keys[key_key] = [key.lower()]
        else:
            keys[key_key].append(key.lower())

    #Returns list of wards and th created key
    return w_list,key

#Generates the districts from the BLUE BOOK file
def gen_districts(bb_wb,districts,keys):

    #Opens ACTIVE page of the BLUE BOOK excell spreadsheet
    bb_ws = bb_wb.active

    #Saves the titlename of the active BLUE BOOK sheet
    global first_sheet
    first_sheet = bb_ws.title

    #Iterates throw rows of the BLUE BOOK, saving individual sheets
    for row in bb_ws:

        #The data begins when this row is not empty. Creates and stores District
        if row[WARD].value:
            temp_d = District(row,keys)
            districts[temp_d.str.lower()] = temp_d

#Collects the assembly voting data from the rest of the BlUE BOOK file
def collect_asmbly_data(bb_wb,districts,keys):

    #Iteratetes through all district sheets in the spreadsheet
    for bb_ws in bb_wb.worksheets:

        #Skip the first
        if bb_ws.title == first_sheet:
            continue

        for row in bb_ws:

            #Check if entry actually exists
            if row[WARD].value:

                #Convert the row into a key value
                temp_key = convert_to_key(row,keys,False)[1].lower()

                #If this key already exists, update that Districts data
                if temp_key in districts.keys():
                    districts[temp_key].dass = row[DEM_PRES].value
                    districts[temp_key].rass = row[REP_PRES].value

def process_keys_info(templist,keys):

    #Separate templist head into key format
    index = [p for p, char in enumerate(templist[0][0].value) if char == ' '][-1] #Last ' ' preceeds numbers
    key = str(templist[0][1].value.replace(" ", "")+" "+templist[0][0].value[:index]).lower() #Cut off numbers, add county
    key_len = len(templist[0][0].value[:index]) #Save length

    #Getting the numbers of the wards from the LTSB data
    templist_nums = []
    for k_l in templist:
        templist_nums.append(k_l[0].value[key_len+1:]) #Append all numbers

    return key,templist_nums

#Exports the list of wards assigned to a town
def export_list(templist,keys,districts):

    key, templist_nums = process_keys_info(templist,keys)
    #Generate list of subliststs in to subdivide templist into ward groupings
    sublists =[]

    if key in keys.keys():

        for key_list in keys[key]: #Retrieve keys matching templist
            print("here")
            short_kl = key_list[len(key):].replace(" ", "").split(",") #Isolate lists of numbers from the keylist

            sublists.append([]) #Start a new sublist to group ward clusters

            #Attempt to match wards from BB to LTSB using ward numbers
            for copy_key in short_kl:

                #Appends the row with the identified ward number
                if copy_key in templist_nums:
                    sublists[-1].append(templist[templist_nums.index(copy_key)])

                else:
                    print("Error: " +str(short_kl) + " " +str(copy_key) +" is not valid for this file")

            #Add the sublist to the proper dictionary
            districts[key_list.lower()].ltsb_append(sublists[-1])


#Parses the data stored in the R file based on what is in the BlUE BOOK. Temporary
#lists of wards contained within a single town are created, then exported into sublists
#divided by ward groupings specified by the BLUE BOOK. WARNING: Heavy string manipulation
def parse_data(r_ws,districts,keys):

    #Creates a list of wards to be broken down using the BlUE BOOK data
    templist = []
    listing = False

    #Iterates through the R worksheet
    for row in r_ws:

        #Skips the first line and makes sure the templist is active
        if row[0].value == "ward":

            global ROW_E

            ROW_E = np.concatenate([District.append_data(HISP_18+1,TOT_18,row,False),
                                    District.append_data(HISP_VEP+1,TOT_VEP,row,False),
                                    District.append_data(DEM_INCUM+1,REP_INCUM,row,False),
                                    District.append_data(WSAREP+2,ECO1,row,False)],0)
            # input(ROW_E[ARR_INCUM])

        elif templist is not None:

            #Starting the list with a new row
            if not listing:
                templist.append(row)
                listing = True

            else:

                #Saving the "ward" information from the R file
                tempval = row[0].value

                #Starts a new value if our temporary list has been exported
                if tempval == None:
                    export_list(templist,keys,districts)
                    templist = None
                    continue

                #Ensures that longer numbers are cut to the correct length
                if len(tempval)>=len(templist[0][0].value):
                    tempval = row[0].value[:len(templist[0][0].value)]

                #If the head of our list contains the same town/county/ward info
                # as the next value in the list, append it
                if tempval[:-1] == templist[0][0].value[:-1]:
                    templist.append(row)

                #Exports the list
                else:
                    export_list(templist,keys,districts)
                    templist = [row]


#Exports all of the data contained in the District files into a new excel spreadsheet
def export_spreadsheet(districts):

    nr_wb = xl.Workbook()
    nr_ws = nr_wb.active

    first_row = np.concatenate([ROW1,ROW_E],0)

    for name in first_row:

        nr_ws.cell(row=1,column=(np.where(name==first_row)[0][0]+1)).value = name
        print(np.where(name==first_row)[0][0],name)

    s_row = 2

    for ID, ward_group in districts.items():

        nr_ws.cell(row=s_row, column=KEY).value = ID
        nr_ws.cell(row=s_row, column=PREP).value = ward_group.rprs
        nr_ws.cell(row=s_row, column=PDEM).value = ward_group.dprs
        nr_ws.cell(row=s_row, column=AREP).value = ward_group.rass
        nr_ws.cell(row=s_row, column=ADEM).value = ward_group.dass
        nr_ws.cell(row=s_row, column=CTY).value = ward_group.cty
        nr_ws.cell(row=s_row, column=ASMB).value = ward_group.asmbly
        nr_ws.cell(row=s_row, column=FIP).value = None

        if ward_group.ltsb is not None:
            if not int(ward_group.ltsb[ARR_INCUM]) ==  0:
                ward_group.ltsb[ARR_INCUM] = 1

            for i in range(len(ward_group.ltsb)):

                nr_ws.cell(row=s_row, column=DEMOGRAPHICS+i).value = ward_group.ltsb[i]


        s_row += 1

    nr_wb.save(output_file)



if __name__ == "__main__":

    bb_wb = xl.load_workbook(bb_file)

    r_wb = xl.load_workbook(r_file)
    r_ws = r_wb.active

    spreadsheet = {}
    districts = {}
    keys = {}

    gen_districts(bb_wb,districts,keys)

    collect_asmbly_data(bb_wb,districts,keys)

    parse_data(r_ws,districts,keys)

    export_spreadsheet(districts)
